from __future__ import annotations

import argparse
import base64
from io import BytesIO
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_MARKERS = {
    "MethodVersion", "SiteLabel", "machine", "patient_key", "field_name",
    "source_name", "service_date", "period_month",
}


def _find_header_row(ws) -> int:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30), values_only=True), 1):
        values = {str(value).strip() for value in row if value is not None}
        if len(values) >= 2 and values.intersection(HEADER_MARKERS):
            return row_number
    raise ValueError(f"Keine Tabellenkopfzeile in Blatt {ws.title!r} gefunden")


def load_ssrs_sheet(source: Path, sheet_name: str) -> pd.DataFrame:
    source = Path(source)
    workbook = load_workbook(source, read_only=False, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Blatt {sheet_name!r} fehlt in {source.name}")
    ws = workbook[sheet_name]
    header_row = _find_header_row(ws)
    values = list(ws.iter_rows(min_row=header_row, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in values[0]]
    keep = [index for index, header in enumerate(headers) if header]
    rows = [[row[index] if index < len(row) else None for index in keep] for row in values[1:]]
    frame = pd.DataFrame(rows, columns=[headers[index] for index in keep])
    return frame.dropna(how="all").reset_index(drop=True)


def _matching_sheets(source: Path, prefix: str) -> list[str]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    return [name for name in workbook.sheetnames if name == prefix or name.startswith(prefix + "_")]


def load_ssrs_sheets(source: Path, prefix: str) -> pd.DataFrame:
    sheets = _matching_sheets(source, prefix)
    if not sheets:
        return pd.DataFrame()
    return pd.concat([load_ssrs_sheet(source, name) for name in sheets], ignore_index=True)


def _numeric(frame: pd.DataFrame, columns: Iterable[str]) -> pd.DataFrame:
    frame = frame.copy()
    for column in columns:
        if column in frame:
            frame[column] = pd.to_numeric(frame[column], errors="coerce")
    return frame


def summarize_waiting_times(source: Path) -> pd.DataFrame:
    details = load_ssrs_sheets(Path(source), "91_Appointments")
    if details.empty or "machine" not in details:
        return pd.DataFrame()
    metrics = [
        "arrival_to_clinical_start_minutes",
        "slot_to_clinical_start_minutes",
        "pending_to_clinical_start_minutes",
        "arrival_to_pending_minutes",
        "clinical_start_to_completed_minutes",
    ]
    details = _numeric(details, metrics)
    rows = []
    for machine, group in details.groupby("machine", dropna=False):
        arrival = group.get("arrival_to_clinical_start_minutes", pd.Series(dtype=float)).dropna()
        slot = group.get("slot_to_clinical_start_minutes", pd.Series(dtype=float)).dropna()
        pending = group.get("pending_to_clinical_start_minutes", pd.Series(dtype=float)).dropna()
        arrival_to_pending = group.get("arrival_to_pending_minutes", pd.Series(dtype=float)).dropna()
        completed = group.get("clinical_start_to_completed_minutes", pd.Series(dtype=float)).dropna()
        rows.append({
            "machine": machine,
            "matched_appointments": int(group.get("matched_session_key", pd.Series(index=group.index)).notna().sum()),
            "arrival_timestamps": int(group.get("patient_arrival_timestamp", pd.Series(index=group.index)).notna().sum()),
            "arrival_samples": int(arrival.size),
            "arrival_after_start": int((group.get("arrival_timing_quality", pd.Series(index=group.index)) == "after_clinical_start").sum()),
            "arrival_earlier_than_6h": int((group.get("arrival_timing_quality", pd.Series(index=group.index)) == "earlier_than_6h").sum()),
            "median_arrival_to_clinical_start_minutes": float(arrival.median()) if not arrival.empty else None,
            "p75_arrival_to_clinical_start_minutes": float(arrival.quantile(0.75)) if not arrival.empty else None,
            "slot_delta_samples": int(slot.size),
            "median_slot_to_clinical_start_minutes": float(slot.median()) if not slot.empty else None,
            "pending_samples": int(pending.size),
            "median_pending_to_clinical_start_minutes": float(pending.median()) if not pending.empty else None,
            "arrival_to_pending_samples": int(arrival_to_pending.size),
            "median_arrival_to_pending_minutes": float(arrival_to_pending.median()) if not arrival_to_pending.empty else None,
            "completed_samples": int(completed.size),
            "median_clinical_start_to_completed_minutes": float(completed.median()) if not completed.empty else None,
        })
    return pd.DataFrame(rows).sort_values("machine").reset_index(drop=True)


def summarize_session_timing(source: Path) -> pd.DataFrame:
    details = load_ssrs_sheets(Path(source), "90_Sessions")
    if details.empty or "machine" not in details:
        return pd.DataFrame()
    details = _numeric(details, ["imaging_to_beam_minutes", "imaging_after_beam_flag"])
    rows = []
    for machine, group in details.groupby("machine", dropna=False):
        imaging = group.get("first_imaging_timestamp", pd.Series(index=group.index)).notna()
        intervals = group.get("imaging_to_beam_minutes", pd.Series(dtype=float)).dropna()
        rows.append({
            "machine": machine,
            "sessions": int(len(group)),
            "imaging_sessions": int(imaging.sum()),
            "imaging_coverage_pct": round(100.0 * float(imaging.sum()) / len(group), 2) if len(group) else None,
            "median_imaging_to_beam_minutes": float(intervals.median()) if not intervals.empty else None,
            "p75_imaging_to_beam_minutes": float(intervals.quantile(0.75)) if not intervals.empty else None,
            "imaging_after_beam": int(group.get("imaging_after_beam_flag", pd.Series(index=group.index)).fillna(0).sum()),
        })
    return pd.DataFrame(rows).sort_values("machine").reset_index(drop=True)


def _machine_summary(source: Path) -> pd.DataFrame:
    days = _numeric(load_ssrs_sheet(source, "03_DeviceDay"), [
        "delivered_sessions", "treated_patients", "gross_device_hours", "net_proxy_hours",
        "median_cycle_minutes", "avg_actual_minutes",
    ])
    if days.empty:
        return days
    return days.groupby("machine", as_index=False).agg(
        delivered_sessions=("delivered_sessions", "sum"),
        patient_contacts=("treated_patients", "sum"),
        active_device_days=("service_date", "nunique"),
        gross_device_hours=("gross_device_hours", "sum"),
        net_proxy_hours=("net_proxy_hours", "sum"),
        median_daily_cycle_minutes=("median_cycle_minutes", "median"),
        median_daily_actual_minutes=("avg_actual_minutes", "median"),
    )


def _monthly_summary(source: Path) -> pd.DataFrame:
    monthly = _numeric(load_ssrs_sheet(source, "02_MachineMonth"), [
        "delivered_sessions", "treated_patients", "active_device_days",
    ])
    if "period_month" in monthly:
        monthly["period_month"] = pd.to_datetime(monthly["period_month"], errors="coerce")
    return monthly


def _slot_summary(source: Path) -> pd.DataFrame:
    slots = _numeric(load_ssrs_sheet(source, "04_SlotSummary"), [
        "relevant_slots", "matched_slots", "match_coverage_pct", "avg_scheduled_minutes",
        "avg_actual_minutes", "avg_slot_utilization_pct",
    ])
    if slots.empty:
        return slots
    return slots.sort_values(["machine", "relevant_slots"], ascending=[True, False]).reset_index(drop=True)


def _site_overview(source: Path) -> pd.DataFrame:
    summary = load_ssrs_sheet(source, "01_SiteSummary")
    if summary.empty:
        return summary
    wanted = [
        "SiteLabel", "requested_period_start", "requested_period_end",
        "observed_first_treatment_date", "observed_last_treatment_date",
        "active_machines", "delivered_sessions", "unique_patients",
        "active_calendar_days", "active_saturdays", "active_device_days", "gross_device_hours",
    ]
    return summary[[column for column in wanted if column in summary.columns]]


def _chart_data_uri(monthly: pd.DataFrame, value: str, title: str, ylabel: str) -> str:
    if monthly.empty or value not in monthly:
        return ""
    import matplotlib.pyplot as plt

    pivot = monthly.pivot_table(index="period_month", columns="machine", values=value, aggfunc="sum").sort_index()
    fig, ax = plt.subplots(figsize=(10, 4.2))
    pivot.plot(ax=ax, marker="o", linewidth=1.8)
    ax.set_title(title)
    ax.set_xlabel("")
    ax.set_ylabel(ylabel)
    ax.grid(axis="y", alpha=0.25)
    fig.tight_layout()
    buffer = BytesIO()
    fig.savefig(buffer, format="png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    return "data:image/png;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")


def _html_table(frame: pd.DataFrame, empty_text: str) -> str:
    if frame.empty:
        return f"<p class='muted'>{empty_text}</p>"
    return frame.to_html(index=False, border=0, classes="data", na_rep="-")


def _write_html(target: Path, title: str, overview: pd.DataFrame, machines: pd.DataFrame,
                monthly: pd.DataFrame, slots: pd.DataFrame, waiting: pd.DataFrame,
                session_timing: pd.DataFrame) -> None:
    chart = _chart_data_uri(monthly, "delivered_sessions", "Behandlungen je Monat und Geraet", "Sitzungen")
    chart_html = f"<img class='chart' src='{chart}' alt='Monatlicher Durchsatz'>" if chart else ""
    target.write_text(f"""<!doctype html>
<html lang="de"><head><meta charset="utf-8"><title>{title}</title><style>
body{{font-family:Arial,sans-serif;margin:32px;color:#172a3a;background:#f5f7fa}}main{{max-width:1280px;margin:auto}}
h1{{margin-bottom:4px}}h2{{margin-top:34px;border-bottom:2px solid #0b6e99;padding-bottom:7px}}
.muted{{color:#607080}}.chart{{width:100%;max-width:1050px;background:white;border:1px solid #d7e0e8}}
table.data{{border-collapse:collapse;width:100%;background:white;font-size:13px}}table.data th,table.data td{{border:1px solid #d7e0e8;padding:7px;text-align:right}}
table.data th:first-child,table.data td:first-child{{text-align:left}}table.data th{{background:#123a63;color:white}}
.note{{background:#eaf3f8;border-left:4px solid #0b6e99;padding:12px}}
</style></head><body><main><h1>{title}</h1><p class="muted">Automatisch aus dem pseudonymisierten ARIA-18-Collector-Export erzeugt.</p>
<div class="note">Der klinische Start verwendet die erste Bildgebung, falls sie dokumentiert ist, sonst den ersten Beam. Beide Originalzeitpunkte bleiben im Collector getrennt erhalten.</div>
<h2>Standort und Zeitraum</h2>{_html_table(overview, 'Keine Standortdaten vorhanden.')}
<h2>Geraete</h2>{_html_table(machines, 'Keine Geraetedaten vorhanden.')}
<h2>Zeitlicher Verlauf</h2>{chart_html}
<h2>Termin- und Wartezeiten</h2>{_html_table(waiting, 'Keine pseudonymisierten Termindetails aktiviert oder keine Ereigniszeitpunkte vorhanden.')}
<h2>Bildgebung und erster Beam</h2>{_html_table(session_timing, 'Keine pseudonymisierten Sitzungsdetails aktiviert oder keine Bildgebungszeitpunkte vorhanden.')}
<h2>Slotnutzung</h2>{_html_table(slots.head(30), 'Keine belastbaren Slotdaten vorhanden.')}
</main></body></html>""", encoding="utf-8")


def _format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="123A63")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(wrap_text=True)
        for column in ws.columns:
            width = min(42, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
            ws.column_dimensions[column[0].column_letter].width = width
    if "Monatsverlauf" in wb.sheetnames:
        ws = wb["Monatsverlauf"]
        if ws.max_row > 2 and ws.max_column >= 3:
            chart = LineChart()
            chart.title = "Sitzungen im Monatsverlauf"
            chart.y_axis.title = "Sitzungen"
            chart.add_data(Reference(ws, min_col=3, max_col=3, min_row=1, max_row=ws.max_row), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
            ws.add_chart(chart, "J2")
    if "Geraete" in wb.sheetnames:
        ws = wb["Geraete"]
        if ws.max_row > 2:
            chart = BarChart()
            chart.title = "Sitzungen je Geraet"
            chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=ws.max_row), titles_from_data=True)
            chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
            ws.add_chart(chart, "J2")
    wb.save(path)


def analyze(source: Path, output_directory: Path, title: str) -> tuple[Path, Path]:
    source = Path(source)
    output_directory = Path(output_directory)
    output_directory.mkdir(parents=True, exist_ok=True)
    overview = _site_overview(source)
    machines = _machine_summary(source)
    monthly = _monthly_summary(source)
    slots = _slot_summary(source)
    waiting = summarize_waiting_times(source)
    session_timing = summarize_session_timing(source)
    quality = load_ssrs_sheet(source, "07_DataQuality")
    glossary = load_ssrs_sheet(source, "08_Glossary") if "08_Glossary" in _matching_sheets(source, "08_Glossary") else pd.DataFrame()

    stem = source.stem + "_Einzelstandortauswertung"
    excel_path = output_directory / f"{stem}.xlsx"
    html_path = output_directory / f"{stem}.html"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Uebersicht", index=False)
        machines.to_excel(writer, sheet_name="Geraete", index=False)
        monthly.to_excel(writer, sheet_name="Monatsverlauf", index=False)
        slots.to_excel(writer, sheet_name="Slots", index=False)
        waiting.to_excel(writer, sheet_name="Wartezeiten", index=False)
        session_timing.to_excel(writer, sheet_name="Imaging_und_Beam", index=False)
        quality.to_excel(writer, sheet_name="Datenqualitaet", index=False)
        glossary.to_excel(writer, sheet_name="Glossar", index=False)
    _format_workbook(excel_path)
    _write_html(html_path, title, overview, machines, monthly, slots, waiting, session_timing)
    return excel_path, html_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Erzeugt eine Einzelstandortauswertung aus einem ARIA-18-Collector-Excel-Export.")
    parser.add_argument("input", type=Path, help="Collector-Exceldatei (.xlsx)")
    parser.add_argument("--output-dir", type=Path, default=Path("analysis_output"))
    parser.add_argument("--title", default="ARIA-18 Durchsatzanalyse - Einzelstandort")
    args = parser.parse_args()
    excel_path, html_path = analyze(args.input, args.output_dir, args.title)
    print(f"ANALYSIS_EXCEL: {excel_path}")
    print(f"ANALYSIS_HTML: {html_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
